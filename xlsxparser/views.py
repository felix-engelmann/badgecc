from django.shortcuts import render, redirect
from django import forms

from openpyxl import load_workbook

from persons.models import Role,Department,Person
from .models import Image

import difflib

import json

class UploadFileForm(forms.Form):
    file = forms.FileField(label='xlsx Dokument')
    #submit = forms.SubmitField(label='hochladen')

# Create your views here.

def index(request):
    #return render(request, "parser/index.html")
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            
            roles=Role.objects.all()
            
            sess={}
            i=0
            
            d={}
            warn_names=[]
            warn_dep=[]
            warn_double=[]
            fail_update=[]
            
            wb = load_workbook(filename = request.FILES['file'].file)
            form= ", ".join(wb.get_sheet_names())
            for sheet in wb.get_sheet_names():
                #if sheet != "Longdrink":
                #pass
                #continue
                
                #get column description
                name_row=0
                fs_row=0
                role_row=0
                mail_row=0
                dep_row=-1
                for idx,head in enumerate(wb[sheet].rows[0]):
                    if not head.value:
                        continue
                    if(head.value.lower()=="name"):
                        name_row=idx
                    if(head.value.lower()=="rolle"):
                        role_row=idx
                    if(head.value.lower()=="fachschaft"):
                        fs_row=idx
                    if(head.value.lower()=="email"):
                        mail_row=idx
                    if(head.value.lower()=="stand"):
                        dep_row=idx
                
                d[sheet]=[]
                for row in wb[sheet].rows[1:]:
                    #might be a valid line
                    email=str(row[mail_row].value)
                    if(row[name_row].value and ( row[fs_row].value or row[role_row].value or ( email.find('@') != -1) ) ):
                        fn=" ".join(row[name_row].value.split(' ')[:-1])
                        ln=row[name_row].value.split(' ')[-1]
                        if dep_row != -1:
                            dep=row[dep_row].value
                            if dep not in d:
                                d[dep]=[]
                        else:
                            dep=sheet
                        
                        correct_role=True
                        role=None
                        try:
                            if(row[role_row].value):
                                role = Role.objects.get(name=row[role_row].value)
                            department=Department.objects.get(name=dep)
                        except:
                            correct_role=False
                        
                        try:
                            if(Person.objects.get(firstname=fn,lastname=ln,role=role,department=department)):
                                warn_double.append({"firstname":fn,"lastname":ln, "role":role, "department":dep})
                                continue
                        except:
                            pass
                        
                        try:
                            op=Person.objects.get(firstname=fn,lastname=ln)
                            fail_update.append({"firstname":fn,"lastname":ln, "role":role,"old_role":op.role, "department":dep,"old_department":op.department})
                        except:
                            pass
                        
                        d[dep].append({"firstname":fn,"lastname":ln,"role":row[role_row].value,"id":i , "guess_role":correct_role})
                        sess[str(i)]={"firstname":fn,"lastname":ln,"role":row[role_row].value,"department":dep}
                        i+=1
                    elif(row[name_row].value):
                        warn_names.append(row[name_row].value)
            request.session["xlsx"]=json.dumps(sess)
            return render(request,"parser/people_select.html",{"data":d,"warnings":warn_names,"roles":roles,"warn_double":warn_double,"fail_update":fail_update})
        #else:
        return render(request,"parser/index.html",{'form': form})
    else:
        form = UploadFileForm()
        return render(request,"parser/index.html",{'form': form})

def insert(request):
    if request.method == 'POST':
        imps=[]
        persons=json.loads(request.session["xlsx"])
        for p_id in request.POST.getlist('insert'):
            p=persons[str(p_id)]
            role=None
            if p['role']:
                role = Role.objects.get(name=p['role'])
            department=Department.objects.get(name=p['department'])
            np=Person(firstname=p['firstname'],lastname=p['lastname'],department=department,role=role)
            np.save()
            imps.append(p)
        return render(request,"parser/person_list.html",{"persons":imps})
    else:
        form = UploadFileForm()
        return render(request,"parser/index.html",{'form': form})

def images(request):
    if request.method == 'POST':
        #form = UploadFileForm(request.POST, request.FILES)
        #if form.is_valid():
            #wb = load_workbook(filename = request.FILES['file'].file)
            #form= ", ".join(wb.get_sheet_names())
        #else:
        imgs=[]
        for f in request.FILES.getlist('files'):
            m=Image()
            m.image = f
            m.original = ".".join(f.name.split('.')[:-1])
            
            
            #filesystem firendly name
            
            osnames={}
            for p in Person.objects.all():
                osnames.update({ (p.firstname+" "+p.lastname).strip().lower().replace(" ","_").replace("ö","o").replace("ä","a").replace("ü","u").replace("ß","s").replace("é","e").replace(".","_") : p.id })
            
            #osnames=list(map( lambda x: (x.firstname+" "+x.lastname).strip().lower().replace(" ","_").replace("ö","o").replace("ä","a").replace("ü","u").replace("ß","s").replace("é","e").replace(".","_") ))
            
            matches=difflib.get_close_matches(m.original,list(osnames.keys()),5,0.98)
    
            person = None
            
            if len(matches)==1:
                person = Person.objects.get(id=osnames[matches[0]])
            
            m.possible = person
            
            m.save()
            
            imgs.append(m)
            
        return render(request,"parser/image_select.html",{"images":imgs,"persons":Person.objects.order_by("lastname")})
    else:
        return render(request,"parser/images.html")

def upload(request):
    if request.method == 'POST':
        for k,v in request.POST.dict().items():
            k = k.split("|")
            if(k[0] == "img" and v!="-1"):
                #print(k[1]+"->"+v)
                pers=Person.objects.get(id=v)
                img=Image.objects.get(id=k[1]).image
                pers.image = img
                pers.save()
                #print(pers.lastname+"->"+img.file.name)
        return redirect('/persons/')
    else:
        return render(request,"parser/images.html")
        
def manual(request):
    pass